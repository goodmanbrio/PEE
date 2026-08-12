#!/usr/bin/env python3
"""
date created: 2026-08-04
date updated: 2026-08-11
date surroundings last checked: 2026-08-11
"""
import json
import math
import os
import sys

import openpyxl
import requests

from dateregex import is_date_like, is_year_match

HARNESS_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(HARNESS_DIR, ".env")
PROTOCOL_PATH = os.path.join(HARNESS_DIR, "BunNav1.md")

DEEPSEEK_BASE_URL = "https://api.deepseek.com/v1/chat/completions"
DEFAULT_MODEL = "deepseek-v4-pro"

MAX_TOOL_CALLS = 40
PEEK_MAX_CELLS = 50
PEEK_ASCII_MAX_CELLS = 225
MINIMAP_BIN = 10
# Row-count budget replaces the old flat ROW_SCAN_CAP/COL_SCAN_CAP pair (wsnSearch4.md
# item 5): measured cost scales with rows scanned, not cell count — column width is
# nearly free even at Excel's near-absolute column ceiling. No column cap.
ROW_SCAN_BUDGET = 10_000
# Payload cap, independent of the scan-depth budget above: how many values all_axis
# returns, not how far it scans. wsnSearch4.md item 5 §1 "wrinkle".
AXIS_VALUE_CAP = 500
# peek_row/peek_col window radius and char safety net — wsnSearch4.md item 2 §2.
PEEK_AXIS_WINDOW_RADIUS = 20
PEEK_AXIS_CHAR_LIMIT = 1000
MAX_TOOL_OUTPUT_CHARS = 6000
BLOCK_MERGE_MAX_GAP = 3
BLOCK_MERGE_MAX_ROWSUM = 20
# map_block's AB-run/CD-run detector — wsnSearch5.md item 2.
RUN_LENGTH_THRESHOLD = 3
CANDIDATE_CAP = 5
SEQ_YEAR_MIN, SEQ_YEAR_MAX = 1980, 2035
# Display formatting only — wsnSearch5.md item 1. Significant figures, not decimal places: a
# fixed decimal-place cap keeps 7 sig figs on a value like 446.08 but only 1-2 on a sub-1 value
# like 0.009499211388800962 (rounds to 0.009, a 5.26% error — enough to fail a 0.5%-tolerance
# grading check even when the exact right cell was read). wsnSearch5_testlog.md q04.
NUMBER_DISPLAY_SIGFIGS = 3

# set_terms (wsnSearch5.md item 4) needs a default distinguishable from an explicit
# None — None is a legitimate value meaning "clear this term".
_UNSET = object()

EXCEL_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA"}

# DeepSeek v4 pro pricing, USD per million tokens (api-docs.deepseek.com/quick_start/pricing)
PRICE_PER_MTOK_INPUT_HIT = 0.003625
PRICE_PER_MTOK_INPUT_MISS = 0.435
PRICE_PER_MTOK_OUTPUT = 0.87


def is_excel_error(v):
    return isinstance(v, str) and v in EXCEL_ERRORS


def _round_display(v):
    """Cap a displayed number cell to NUMBER_DISPLAY_SIGFIGS significant figures — wsnSearch5.md
    item 1. Prevents payload bloat on cells like 47.83973533973534, and (unlike a fixed
    decimal-place cap) stays accurate on sub-1 values like 0.009499211388800962. Ints/strings/
    None pass through unchanged; 0.0 has no sig-fig position and passes through as-is."""
    if not isinstance(v, float):
        return v
    if v == 0.0:
        return v
    digits = NUMBER_DISPLAY_SIGFIGS - int(math.floor(math.log10(abs(v)))) - 1
    return round(v, digits)


def _resolve_term(explicit, stored):
    """Session-term fallback for map_block/peek_ascii/list_sheets/map_bin — wsnSearch5.md
    item 4. Omitted at call time (_UNSET) falls back to the set_terms default; an explicit
    value (including None) applies to this call only and doesn't touch the stored default."""
    return stored if explicit is _UNSET else explicit


def _parse_range(s):
    """'3-6' -> (3, 6); '3' -> (3, 3). rows/cols block-range strings, always copied verbatim
    off a map_block output row — check_axes/peek_axes item 3a/3b contract."""
    s = str(s).strip()
    if "-" in s:
        lo, hi = s.split("-", 1)
        return int(lo), int(hi)
    v = int(s)
    return v, v


def _parse_tag(t):
    """'3r' -> (3, 'r'); '7c' -> (7, 'c'). Tags are always copied verbatim off a map_block
    AB-run/CD-run column — check_axes/peek_axes item 3a/3b contract."""
    t = t.strip()
    orient = t[-1].lower()
    if orient not in ("r", "c"):
        raise ValueError(f"tag '{t}' doesn't end in 'r' or 'c'")
    return int(t[:-1]), orient


def _matches(term, v):
    return bool(term) and isinstance(v, str) and term.lower() in v.lower()


def _year_matches(term, v):
    return term is not None and is_year_match(v, term)


def _cell_flags(v, metric_term=None, entity_term=None, year_term=None):
    """Independent (date, metric, entity, year) booleans for one cell — not priority-
    collapsed. A cell can be true for more than one at once (e.g. "300mm wafer" matches
    metric and entity; a bare "2022" matches date and year)."""
    if v is None:
        return (False, False, False, False)
    return (is_date_like(v), _matches(metric_term, v), _matches(entity_term, v), _year_matches(year_term, v))


def _cell_code(v, metric_term=None, entity_term=None, year_term=None):
    """Single-letter priority code for one cell: blank > metric > entity > year > date >
    string/number/other."""
    if v is None:
        return "."
    if _matches(metric_term, v):
        return "A"
    if _matches(entity_term, v):
        return "B"
    if _year_matches(year_term, v):
        return "C"
    if is_date_like(v):
        return "D"
    if isinstance(v, str):
        return "S"
    if isinstance(v, (int, float)):
        return "N"
    return "?"


def _render_grid(symbol_rows, row_labels, col_labels):
    row_label_strs = [str(x) for x in row_labels]
    col_label_strs = [str(x) for x in col_labels]
    row_w = max((len(s) for s in row_label_strs), default=1)
    col_w = max((len(s) for s in col_label_strs), default=1)
    header = " " * row_w + " " + " ".join(s.rjust(col_w) for s in col_label_strs)
    lines = [header]
    for rlabel, srow in zip(row_label_strs, symbol_rows):
        lines.append(rlabel.rjust(row_w) + " " + " ".join(s.rjust(col_w) for s in srow))
    return "\n".join(lines)


def _render_grid_var(symbol_rows, row_labels, col_labels):
    """Like _render_grid, but each column's width comes from its own content, not just the
    label — check_axes can render literal cell values of arbitrary width, including
    multi-word strings with internal spaces, which a label-only column width would misalign.
    peek_ascii/map_bin don't need this (their content is always 1 or 5 chars, already <= any
    real label width) and keep using _render_grid unchanged."""
    row_label_strs = [str(x) for x in row_labels]
    col_label_strs = [str(x) for x in col_labels]
    row_w = max((len(s) for s in row_label_strs), default=1)
    col_w = []
    for ci, clabel in enumerate(col_label_strs):
        w = len(clabel)
        for srow in symbol_rows:
            w = max(w, len(str(srow[ci])))
        col_w.append(w)
    header = " " * row_w + " " + " ".join(s.rjust(w) for s, w in zip(col_label_strs, col_w))
    lines = [header]
    for rlabel, srow in zip(row_label_strs, symbol_rows):
        lines.append(rlabel.rjust(row_w) + " " + " ".join(str(s).rjust(w) for s, w in zip(srow, col_w)))
    return "\n".join(lines)


def _render_table(headers, rows):
    cols = [headers] + rows
    widths = [max(len(str(row[i])) for row in cols) for i in range(len(headers))]
    def fmt(row):
        return "  ".join(str(x).ljust(w) for x, w in zip(row, widths))
    return "\n".join(fmt(r) for r in cols)


def _detect_blocks(values, n_rows, n_cols):
    """Column bands split by fully-blank columns, then row bands (blocks) split by fully-blank
    rows within each band. Returns zones in detection order (band-major, top-to-bottom), each
    tagged with its band id so callers can enforce "never merge across column bands"."""
    blank_col = [all(values[r][c] is None for r in range(n_rows)) for c in range(n_cols)]

    col_bands = []
    c = 0
    while c < n_cols:
        if blank_col[c]:
            c += 1
            continue
        start = c
        while c < n_cols and not blank_col[c]:
            c += 1
        col_bands.append((start + 1, c))

    zones = []
    band_id = 0
    for (c0, c1) in col_bands:
        band_id += 1
        r = 0
        while r < n_rows:
            row_blank = all(values[r][cc - 1] is None for cc in range(c0, c1 + 1))
            if row_blank:
                r += 1
                continue
            start = r
            while r < n_rows and not all(values[r][cc - 1] is None for cc in range(c0, c1 + 1)):
                r += 1
            zones.append({"rows": [start + 1, r], "cols": [c0, c1], "band": band_id})
    return zones


def _merge_blocks(zones):
    """Greedy chain merge within a column band: absorb the next block if separated by 1-3 blank
    rows and the chain's cumulative row-count would stay under 20. Never crosses band boundaries."""
    entries = []
    i = 0
    n = len(zones)
    while i < n:
        nums = [i + 1]
        band = zones[i]["band"]
        start_row, end_row = zones[i]["rows"]
        cols = zones[i]["cols"]
        cum = end_row - start_row + 1
        j = i + 1
        while j < n and zones[j]["band"] == band:
            gap = zones[j]["rows"][0] - end_row - 1
            next_count = zones[j]["rows"][1] - zones[j]["rows"][0] + 1
            if 1 <= gap <= BLOCK_MERGE_MAX_GAP and cum + next_count < BLOCK_MERGE_MAX_ROWSUM:
                nums.append(j + 1)
                end_row = zones[j]["rows"][1]
                cum += next_count
                j += 1
            else:
                break
        entries.append({"nums": nums, "rows": [start_row, end_row], "cols": cols})
        i = j
    return entries


def _find_entry(entries, row, col):
    for e in entries:
        r0, r1 = e["rows"]
        c0, c1 = e["cols"]
        if r0 <= row <= r1 and c0 <= col <= c1:
            return e
    return None


def _is_plain_string(v):
    return isinstance(v, str) and v.strip() != ""


def _is_plain_int_year(v):
    """A plain int/float-that's-really-an-int in SEQ_YEAR_MIN..MAX — candidate cell for the
    sequential-year fallback below. Excel bools are ints in Python; excluded explicitly."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)) and float(v).is_integer():
        iv = int(v)
        if SEQ_YEAR_MIN <= iv <= SEQ_YEAR_MAX:
            return iv
    return None


def _seq_hits_1d(seq):
    """Indices into seq that are part of a maximal run (length >= RUN_LENGTH_THRESHOLD) of
    consecutive integers each exactly +1 from the previous, within SEQ_YEAR_MIN..MAX — the
    numeric-typed-year header case (2012 2013 2014 ...), invisible to both is_date_like (no
    string regex applies to a plain int) and to a fixed year_term match (matches only one
    cell). Feeds CD-run candidate detection only (wsnSearch5.md item 2) — never touches
    is_date_like itself or the D/A/B/C tally counts."""
    hits = set()
    run_start = None
    prev_val = None
    for i, v in enumerate(seq):
        iv = _is_plain_int_year(v)
        if iv is not None and prev_val is not None and iv == prev_val + 1:
            pass  # continues the run already tracked by run_start
        else:
            if run_start is not None and i - run_start >= RUN_LENGTH_THRESHOLD:
                hits.update(range(run_start, i))
            run_start = i if iv is not None else None
        prev_val = iv
    if run_start is not None and len(seq) - run_start >= RUN_LENGTH_THRESHOLD:
        hits.update(range(run_start, len(seq)))
    return hits


def _precompute_seq_grids(values, r0, r1, c0, c1):
    """seq_row[(r,cc)]/seq_col[(r,cc)] = True if that cell is part of a qualifying row-wise /
    col-wise sequential-year run — orientation-specific, one fused pass per orientation, no
    per-cell re-read (wsnSearch5.md item 2 §2, Tier 2 of the CD-run check)."""
    seq_row = {}
    for r in range(r0, r1 + 1):
        row_vals = [values[r - 1][cc - 1] for cc in range(c0, c1 + 1)]
        for i in _seq_hits_1d(row_vals):
            seq_row[(r, c0 + i)] = True
    seq_col = {}
    for cc in range(c0, c1 + 1):
        col_vals = [values[r - 1][cc - 1] for r in range(r0, r1 + 1)]
        for i in _seq_hits_1d(col_vals):
            seq_col[(r0 + i, cc)] = True
    return seq_row, seq_col


def _ab_cd_run_candidates(values, r0, r1, c0, c1, metric_term, entity_term, year_term):
    """map_block's run-detector (wsnSearch5.md item 2). AB-run (lineitem axis): a cell
    continues the run if it's a metric/entity match or a plain non-blank string; the run
    qualifies only if it also contains >=1 real metric/entity hit. CD-run (period axis): a
    cell continues the run if it's date-like (incl. the sequential-int fallback) or a plain
    non-blank string; qualifies only if it contains >=1 real date-like hit. Both computed
    per column and per row (an axis can be found as either), pooled into one candidate list
    per axis type. Returns (ab_candidates, cd_candidates), each a sorted list of
    (index, 'r'|'c') tuples. Never alters the D/A/B/C tally counts — those stay exactly as
    map_block computes them today."""
    seq_row, seq_col = _precompute_seq_grids(values, r0, r1, c0, c1)

    def ab_hit(v):
        _, metric, entity, _ = _cell_flags(v, metric_term, entity_term, year_term)
        anchor = metric or entity
        return (anchor or _is_plain_string(v)), anchor

    def cd_hit(v, r, cc, row_wise):
        seq_anchor = seq_row.get((r, cc), False) if row_wise else seq_col.get((r, cc), False)
        anchor = is_date_like(v) or seq_anchor
        return (anchor or _is_plain_string(v)), anchor

    ab_col, cd_col, ab_row, cd_row = set(), set(), set(), set()

    def _run_pass(outer_range, inner_range, row_wise, out_ab, out_cd):
        for outer in outer_range:
            ab_streak = ab_anchor_ct = cd_streak = cd_anchor_ct = 0
            for inner in inner_range:
                r, cc = (outer, inner) if row_wise else (inner, outer)
                v = values[r - 1][cc - 1]
                hit, anchor = ab_hit(v)
                if hit:
                    ab_streak += 1
                    ab_anchor_ct += anchor
                else:
                    if ab_streak >= RUN_LENGTH_THRESHOLD and ab_anchor_ct >= 1:
                        out_ab.add(outer)
                    ab_streak = ab_anchor_ct = 0
                hit, anchor = cd_hit(v, r, cc, row_wise)
                if hit:
                    cd_streak += 1
                    cd_anchor_ct += anchor
                else:
                    if cd_streak >= RUN_LENGTH_THRESHOLD and cd_anchor_ct >= 1:
                        out_cd.add(outer)
                    cd_streak = cd_anchor_ct = 0
            if ab_streak >= RUN_LENGTH_THRESHOLD and ab_anchor_ct >= 1:
                out_ab.add(outer)
            if cd_streak >= RUN_LENGTH_THRESHOLD and cd_anchor_ct >= 1:
                out_cd.add(outer)

    _run_pass(range(c0, c1 + 1), range(r0, r1 + 1), False, ab_col, cd_col)
    _run_pass(range(r0, r1 + 1), range(c0, c1 + 1), True, ab_row, cd_row)

    ab_candidates = sorted([(c, "c") for c in ab_col] + [(r, "r") for r in ab_row])
    cd_candidates = sorted([(c, "c") for c in cd_col] + [(r, "r") for r in cd_row])
    return ab_candidates, cd_candidates


def _fmt_run_candidates(cands, block_num_str, axis_label, notes_out):
    """Format one axis-type's (AB-run or CD-run) candidate list into map_block's cell text —
    <index><r|c>, comma-joined, '-' if none, capped at CANDIDATE_CAP shown. Appends an
    overflow note (or, past 2x the cap, a stronger 'investigate' note) and/or a mixed-
    orientation note to notes_out — wsnSearch5.md item 2 §2 step 4."""
    if not cands:
        return "-"
    shown = cands[:CANDIDATE_CAP]
    s = ",".join(f"{idx}{orient}" for idx, orient in shown)
    parts = []
    if len(cands) > 2 * CANDIDATE_CAP:
        parts.append(
            f"block {block_num_str} {axis_label} candidate count ({len(cands)}) far exceeds "
            f"typical header-row counts — investigate, could be false positives"
        )
    elif len(cands) > CANDIDATE_CAP:
        parts.append(f"block {block_num_str} {axis_label} {len(cands) - CANDIDATE_CAP} more candidates not shown")
    if len({o for _, o in shown}) > 1:
        parts.append(
            f"block {block_num_str} {axis_label} candidates mix orientation ({s}) — check both, "
            f"one is likely a false positive, order is not a nesting signal"
        )
    if parts:
        notes_out.append("; ".join(parts))
    return s


def _trim_to_char_limit(values_out, anchor_index, char_limit):
    """Char safety net for peek_row/peek_col, independent of the span-based windowing —
    fires even inside a <=41 block whose values happen to be long (wsnSearch4.md item 2
    §2 step 6). Keeps the entries closest to the anchor, drops the rest."""
    rendered = repr(values_out)
    if len(rendered) <= char_limit:
        return values_out, None
    order = sorted(values_out.keys(), key=lambda k: abs(k - anchor_index))
    trimmed = {}
    total = 0
    for k in order:
        piece_len = len(repr({k: values_out[k]}))
        if total + piece_len > char_limit:
            break
        trimmed[k] = values_out[k]
        total += piece_len
    if not trimmed:
        return trimmed, "showing none (single value exceeds char limit)"
    shown = sorted(trimmed.keys())
    return trimmed, f"showing {shown[0]}-{shown[-1]}"


def _clamp_window(r0, r1, c0, c1, cap):
    """Clamp-on-overage for peek/peek_ascii (wsnSearch4.md item 1). Returns (r1, c1, note) —
    r1/c1 unchanged and note=None when already within cap. Tie-break: rows==cols treats
    cols as the axis to clamp."""
    rows = r1 - r0 + 1
    cols = c1 - c0 + 1
    if rows * cols <= cap:
        return r1, c1, None
    cols_is_bigger = cols >= rows
    bigger = cols if cols_is_bigger else rows
    smaller = rows if cols_is_bigger else cols
    if smaller <= cap:
        new_bigger = cap // smaller
        if cols_is_bigger:
            r1_new, c1_new = r1, c0 + new_bigger - 1
        else:
            r1_new, c1_new = r0 + new_bigger - 1, c1
    else:
        # Degenerate: both axes individually exceed cap. A small square samples both
        # dimensions instead of discarding one entirely.
        side = math.isqrt(cap)
        r1_new, c1_new = r0 + side - 1, c0 + side - 1
    return r1_new, c1_new, f"too many cells! truncated to r1={r1_new}, c1={c1_new}"


TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "list_sheets",
            "description": (
                "List visible sheet names (hidden sheets are not accessible — same as what an "
                "analyst opening this file would see by default). If metric_term/entity_term/"
                "year_term are given, also scans every visible sheet and reports raw D/A/B/C "
                "counts and scanned rows/cols per sheet, so you can rank candidate sheets before "
                "opening any in detail. Raw counts only, not a ranking score — a sheet can score "
                "highest simply because the term is its own subject, not because it's the answer; "
                "weigh count against sheet size yourself."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "metric_term": {"type": "string", "description": "optional fuzzy metric term to count matches for (adds an A count per sheet). Omit to use the session default set via set_terms."},
                    "entity_term": {"type": "string", "description": "optional fuzzy entity term to count matches for (adds a B count per sheet). Omit to use the session default set via set_terms."},
                    "year_term": {"type": "integer", "description": "optional target year to count matches for (adds a C count per sheet) — anchor to check, not a settled answer. Omit to use the session default set via set_terms."},
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_dims",
            "description": "Get max_row/max_col for a sheet. Upper bound only, not the real extent — confirm with map_block or map_bin.",
            "parameters": {
                "type": "object",
                "properties": {"sheet": {"type": "string"}},
                "required": ["sheet"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "map_block",
            "description": (
                "Default first structural call after list_sheets. Finds every table block on a "
                "sheet (split by blank columns/rows"
                "within each band) and renders them as a compact line list: one entry per block, "
                "or cluster of small blocks. Each entry gives its row/col range, raw D/A/B/C counts "
                "within, and two more columns, AB-run/CD-run: any unbroken run (length >= 3) of "
                "lineitem-hits (AB) or period-hits (CD) found in that block, as <index><r|c> "
                "candidates (comma-joined, '-' if none). If a block shows a candidate in both "
                "columns, call check_axes directly with those tags/rows/cols — skip blind peek_ascii "
                "sweeps. If only one column has a candidate, the other axis is often stated once in "
                "an earlier block sharing the same cols range, not absent from the sheet. "
                "Call before map_bin — only fall back to map_bin if this flat list doesn't resolve a "
                "spatial/relative-position question (role-swap headers, transposed axes, adjacent "
                "column-band comparison)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "metric_term": {"type": "string", "description": "optional fuzzy metric term; adds an A count per block. Omit to use the session default set via set_terms."},
                    "entity_term": {"type": "string", "description": "optional fuzzy entity term; adds a B count per block. Omit to use the session default set via set_terms."},
                    "year_term": {"type": "integer", "description": "optional target year; adds a C count per block — anchor to check, not a settled answer. Omit to use the session default set via set_terms."},
                },
                "required": ["sheet"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "map_bin",
            "description": (
                "Coarse whole-sheet view, called after map_block when its flat list isn't enough "
                "for spatial/relative-position reasoning (role-swap headers, transposed axes, "
                "adjacent column-band comparison) — a grid conveys that better than a line list. "
                "Returns an ascii minimap binned 10x10 cells per bin. Each bin renders as a "
                "5-character code [structural][D][A][B][C]: structural is '.' blank bin, ':' content "
                "not a corner, '#' a block's exact top-left corner; D/A/B/C are '-' or the letter if "
                "a date-like/metric-match/entity-match/year-match cell exists anywhere in that bin. "
                "A blank bin is always '.----'. Real row/col numbers label each bin boundary — use "
                "them directly as coordinates into peek_ascii."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "metric_term": {"type": "string", "description": "optional fuzzy metric term; sets the A flag on bins containing a match. Omit to use the session default set via set_terms."},
                    "entity_term": {"type": "string", "description": "optional fuzzy entity term; sets the B flag on bins containing a match. Omit to use the session default set via set_terms."},
                    "year_term": {"type": "integer", "description": "optional target year; sets the C flag on bins containing a match — anchor to check, not a settled answer. Omit to use the session default set via set_terms."},
                },
                "required": ["sheet"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "peek_ascii",
            "description": (
                "Render the cell-type grid (not real values) for a bounded window at full resolution, "
                "no binning. One letter per cell, priority order: '.' blank, 'A' metric fuzzy-match, "
                "'B' entity fuzzy-match, 'C' year match, 'D' date-like (regex match or native datetime "
                "dtype), 'S' string, 'N' number, '?' other — query-relevance outranks generic type "
                "info. Real row/col numbers label the grid. Use this to narrow in on a map_block/"
                "map_bin flag before reading real values with peek. Combined cap: "
                f"(r1-r0+1)*(c1-c0+1) should be <= {PEEK_ASCII_MAX_CELLS} — over that, the window is "
                "clamped (bigger axis cut first) and the response carries a 'note' naming the actual "
                "range shown, not rejected."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "r0": {"type": "integer"},
                    "r1": {"type": "integer"},
                    "c0": {"type": "integer"},
                    "c1": {"type": "integer"},
                    "metric_term": {"type": "string", "description": "optional fuzzy metric term; renders matching cells as 'A'. Omit to use the session default set via set_terms."},
                    "entity_term": {"type": "string", "description": "optional fuzzy entity term; renders matching cells as 'B'. Omit to use the session default set via set_terms."},
                    "year_term": {"type": "integer", "description": "optional target year; renders matching cells as 'C' — anchor to check, not a settled answer. Omit to use the session default set via set_terms."},
                },
                "required": ["sheet", "r0", "r1", "c0", "c1"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "peek",
            "description": (
                "Read real values in a small window. Returns {\"rows\": {row_number: {col_number: "
                "value}}} — every value is keyed by its real row/col number, not array position; "
                f"don't count list entries to find a column. Combined cap: (r1-r0+1)*(c1-c0+1) should "
                f"be <= {PEEK_MAX_CELLS} — over that, the window is clamped (bigger axis cut first) "
                "and the response carries a 'note' naming the actual range shown, not rejected."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "r0": {"type": "integer"},
                    "r1": {"type": "integer"},
                    "c0": {"type": "integer"},
                    "c1": {"type": "integer"},
                },
                "required": ["sheet", "r0", "r1", "c0", "c1"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "peek_col",
            "description": (
                "Confirm a lineitem/column placement against its neighbors without dumping a whole "
                "axis: anchor on one cell, get every value in that cell's column across its "
                "containing table block's row range (block found the same way map_block finds one — "
                "no need to call map_block first). Blocks <=41 rows return in full; bigger blocks "
                "window +/-20 rows around the anchor; no block detected falls back to the same "
                "+/-20 window. Sparse output ({\"values\": {row_number: value}}, blanks omitted) plus "
                "a 'note' whenever the window was narrowed, the anchor's own column is blank "
                "throughout, or a further char-length safety net trimmed it."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "row": {"type": "integer", "description": "anchor cell's row"},
                    "col": {"type": "integer", "description": "anchor cell's column — the column being confirmed"},
                },
                "required": ["sheet", "row", "col"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "peek_row",
            "description": (
                "Confirm a period/header placement against its neighbors without dumping a whole "
                "axis: anchor on one cell, get every value in that cell's row across its containing "
                "table block's column range (block found the same way map_block finds one — no need "
                "to call map_block first). Blocks <=41 cols return in full; bigger blocks window "
                "+/-20 cols around the anchor; no block detected falls back to the same +/-20 "
                "window. Sparse output ({\"values\": {col_number: value}}, blanks omitted) plus a "
                "'note' whenever the window was narrowed, the anchor's own row is blank throughout, "
                "or a further char-length safety net trimmed it."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "row": {"type": "integer", "description": "anchor cell's row — the row being confirmed"},
                    "col": {"type": "integer", "description": "anchor cell's column"},
                },
                "required": ["sheet", "row", "col"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "all_axis",
            "description": (
                "Dump every non-blank value along one full row or column — the deliberate full-dump "
                "option for when the whole axis is genuinely needed (prefer peek_row/peek_col for a "
                "targeted neighbor check). kind is 'row' or 'col'. kind='row' scans the full true "
                "column count (column width is cheap regardless of size); kind='col' scans up to "
                f"{ROW_SCAN_BUDGET} rows. Returned values are capped at {AXIS_VALUE_CAP} regardless "
                "of kind — a 'note' names whichever constraint (scan depth or payload size) actually "
                "limited the response."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "kind": {"type": "string", "enum": ["row", "col"]},
                    "index": {"type": "integer"},
                },
                "required": ["sheet", "kind", "index"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_cell",
            "description": "Read a single cell value.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "row": {"type": "integer"},
                    "col": {"type": "integer"},
                },
                "required": ["sheet", "row", "col"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "set_terms",
            "description": (
                "Fix metric_term/entity_term/year_term once for the rest of this session — "
                "list_sheets/map_block/map_bin/peek_ascii all fall back to these whenever you omit "
                "their own term args, so you stop retyping the same strings on every call. Omitting "
                "an arg here leaves it unchanged; passing an explicit null clears it. Returns the "
                "full current state (all three, not just what you touched) after applying the call. "
                "Call this before your first check_axes call — check_axes has no term args of its "
                "own, it only reads whatever set_terms has set."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "metric_term": {"type": "string", "description": "fuzzy metric term for the rest of this session"},
                    "entity_term": {"type": "string", "description": "fuzzy entity term for the rest of this session"},
                    "year_term": {"type": "integer", "description": "target year for the rest of this session"},
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "check_axes",
            "description": (
                "Confirm a map_block AB-run/CD-run candidate is a real axis before reading its "
                "values. sheet/tags/rows/cols are copied verbatim off one map_block output row — no "
                "independent block re-detection. Renders the block (tag-centered, expanded by 20 in "
                "every direction, clamped only if still over "
                f"{PEEK_ASCII_MAX_CELLS} cells): any cell in a tagged row or column shows its real "
                "value (capped to 3 significant figures), every other cell shows peek_ascii's type flag ('.' blank "
                "'A' metric-match 'B' entity-match 'C' year-match 'D' date-like 'S' string 'N' "
                "number '?' other). Real row/col numbers label the grid — but on a wide grid, don't "
                "count grid columns by eye to find a value: the response's separate 'tagged_values' "
                "field restates every tagged row/column's own values as an explicit "
                "{column_or_row_index: value} dict, so read the value there, not off the grid. The "
                "grid itself is for the pattern question (does real data actually line up under both "
                "axes), not for reading exact values. Also auto-includes the row "
                "immediately above a tagged row (col immediately left of a tagged col) when it looks "
                "like a header, even though map_block never flagged it — a coarser nested header "
                "(year sitting above quarter) is often too sparse for map_block's run detector to "
                "catch on its own. Uses whatever metric_term/entity_term/year_term set_terms has set "
                "— call set_terms first, this tool has no term args of its own."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "candidate tags copied from map_block's AB-run/CD-run columns, e.g. [\"3r\",\"7c\"]",
                    },
                    "rows": {"type": "string", "description": "block's row range copied from map_block, e.g. \"3-6\" or \"3\""},
                    "cols": {"type": "string", "description": "block's col range copied from map_block, e.g. \"2-7\" or \"2\""},
                },
                "required": ["sheet", "tags", "rows", "cols"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "peek_axes",
            "description": (
                "Read a confirmed candidate axis's own real values — no grid, no flagging. "
                "sheet/tags/rows/cols are copied verbatim off a map_block row, same as check_axes, "
                "but here rows/cols only bound the scan range per tag, they don't define a rendered "
                "window. For each tag, scans its full line (the tagged row across cols, or the "
                "tagged column across rows) and returns non-blank values only as {index: value} "
                "(3 significant figures), one line per tag, each with its own independent truncation cap — "
                "nothing auto-added beyond what you asked for. Use once a candidate's confirmed (via "
                "check_axes, or trusted directly when map_block returned only one clean candidate) — "
                "cheaper than check_axes for a huge block, since a row's/column's own real values "
                "stay small even when the block itself is huge, and this tool never renders a grid."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "candidate tags copied from map_block's AB-run/CD-run columns, e.g. [\"3r\",\"4r\"]",
                    },
                    "rows": {"type": "string", "description": "scan-range bound copied from map_block, e.g. \"3-14\" or \"3\""},
                    "cols": {"type": "string", "description": "scan-range bound copied from map_block, e.g. \"3-18\" or \"3\""},
                },
                "required": ["sheet", "tags", "rows", "cols"],
            },
        },
    },
]


def load_env(path):
    env = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            value = value.strip()
            if len(value) >= 2 and value[0] == value[-1] and value[0] in "\"'":
                value = value[1:-1]
            env[key.strip()] = value
    return env


class Workbook:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # Session-scoped query terms — wsnSearch5.md item 4. None until set_terms is called,
        # identical to today's behavior for any term-consuming tool called before it.
        self._metric_term = None
        self._entity_term = None
        self._year_term = None

    def set_terms(self, metric_term=_UNSET, entity_term=_UNSET, year_term=_UNSET):
        if metric_term is not _UNSET:
            self._metric_term = metric_term
        if entity_term is not _UNSET:
            self._entity_term = entity_term
        if year_term is not _UNSET:
            self._year_term = year_term
        return {"metric_term": self._metric_term, "entity_term": self._entity_term, "year_term": self._year_term}

    def _check_visible(self, sheet):
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"no sheet named '{sheet}'")
        if self.wb[sheet].sheet_state != "visible":
            raise ValueError(f"sheet '{sheet}' is hidden and not accessible — same as what an analyst opening this file would see")

    def _scan(self, sheet, row_budget=ROW_SCAN_BUDGET):
        ws = self.wb[sheet]
        true_max_row = ws.max_row
        true_max_col = ws.max_column
        max_row = min(true_max_row, row_budget)
        max_col = true_max_col  # no column cap — column width is nearly free to scan
        values = [list(r) for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)]
        n_rows = len(values)
        n_cols = len(values[0]) if n_rows else 0
        return values, n_rows, n_cols, true_max_row, true_max_col

    def list_sheets(self, metric_term=_UNSET, entity_term=_UNSET, year_term=_UNSET):
        metric_term = _resolve_term(metric_term, self._metric_term)
        entity_term = _resolve_term(entity_term, self._entity_term)
        year_term = _resolve_term(year_term, self._year_term)
        results = []
        clamped_sheets = []
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            if ws.sheet_state != "visible":
                continue
            values, n_rows, n_cols, true_max_row, true_max_col = self._scan(name)
            entry = {"name": name, "scanned_rows": n_rows, "scanned_cols": n_cols}
            d = a = b = c = 0
            for row in values:
                for v in row:
                    date, metric, entity, year = _cell_flags(v, metric_term, entity_term, year_term)
                    d += date
                    a += metric
                    b += entity
                    c += year
            entry["D"] = d
            if metric_term is not None:
                entry["A"] = a
            if entity_term is not None:
                entry["B"] = b
            if year_term is not None:
                entry["C"] = c
            if true_max_row > n_rows:
                entry["note"] = f"{name} not fully scanned, {true_max_row - n_rows} rows remaining"
                clamped_sheets.append(name)
            results.append(entry)
        output = {"sheets": results}
        if clamped_sheets:
            output["sheets_not_fully_scanned"] = clamped_sheets
        return output

    def get_dims(self, sheet):
        self._check_visible(sheet)
        ws = self.wb[sheet]
        return {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "note": "upper bound only; formatting artifacts often overstate real extent — confirm with map_block or map_bin",
        }

    def map_block(self, sheet, metric_term=_UNSET, entity_term=_UNSET, year_term=_UNSET):
        metric_term = _resolve_term(metric_term, self._metric_term)
        entity_term = _resolve_term(entity_term, self._entity_term)
        year_term = _resolve_term(year_term, self._year_term)
        self._check_visible(sheet)
        values, n_rows, n_cols, true_max_row, true_max_col = self._scan(sheet)
        if not n_cols:
            return {"blocks": "", "note": "sheet appears empty within scan bounds"}

        zones = _detect_blocks(values, n_rows, n_cols)
        entries = _merge_blocks(zones)

        rows_out = []
        block_notes = []
        for e in entries:
            num_str = str(e["nums"][0]) if len(e["nums"]) == 1 else f"{e['nums'][0]}-{e['nums'][-1]}"
            r0, r1 = e["rows"]
            c0, c1 = e["cols"]
            d = a = b = c = 0
            for r in range(r0, r1 + 1):
                for cc in range(c0, c1 + 1):
                    date, metric, entity, year = _cell_flags(values[r - 1][cc - 1], metric_term, entity_term, year_term)
                    d += date
                    a += metric
                    b += entity
                    c += year

            ab_cands, cd_cands = _ab_cd_run_candidates(values, r0, r1, c0, c1, metric_term, entity_term, year_term)
            ab_str = _fmt_run_candidates(ab_cands, num_str, "AB-run", block_notes)
            cd_str = _fmt_run_candidates(cd_cands, num_str, "CD-run", block_notes)

            rows_out.append([
                num_str,
                str(r0) if r0 == r1 else f"{r0}-{r1}",
                str(c0) if c0 == c1 else f"{c0}-{c1}",
                d, a, b, c, ab_str, cd_str,
            ])

        table = _render_table(["block", "rows", "cols", "D", "A", "B", "C", "AB-run", "CD-run"], rows_out)
        result = {
            "blocks": table,
            "legend": (
                "D date-like count  A metric-match count  B entity-match count  C year-match count — "
                "raw counts over each entry's true row/col extent. AB-run/CD-run: <index><r|c> of any "
                "unbroken run (length >= 3) of lineitem-hits (AB) / period-hits (CD) found in that "
                "block, comma-joined, '-' if none. A block with a candidate in one column but '-' in "
                "the other often has the missing axis stated once in an earlier block sharing the "
                "same cols range, not absent from the sheet."
            ),
            "scanned_rows": n_rows,
            "scanned_cols": n_cols,
        }
        if block_notes:
            result["notes"] = block_notes
        if true_max_row > n_rows:
            result["note"] = f"{sheet} not fully scanned, {true_max_row - n_rows} rows remaining"
        return result

    def map_bin(self, sheet, metric_term=_UNSET, entity_term=_UNSET, year_term=_UNSET):
        metric_term = _resolve_term(metric_term, self._metric_term)
        entity_term = _resolve_term(entity_term, self._entity_term)
        year_term = _resolve_term(year_term, self._year_term)
        self._check_visible(sheet)
        values, n_rows, n_cols, true_max_row, true_max_col = self._scan(sheet)
        if not n_cols:
            return {"minimap": "", "note": "sheet appears empty within scan bounds"}

        zones = _detect_blocks(values, n_rows, n_cols)
        corner_bins = {((z["rows"][0] - 1) // MINIMAP_BIN, (z["cols"][0] - 1) // MINIMAP_BIN) for z in zones}

        n_bin_rows = (n_rows + MINIMAP_BIN - 1) // MINIMAP_BIN
        n_bin_cols = (n_cols + MINIMAP_BIN - 1) // MINIMAP_BIN

        symbol_rows = []
        row_labels = []
        for br in range(n_bin_rows):
            r0, r1 = br * MINIMAP_BIN, min(br * MINIMAP_BIN + MINIMAP_BIN, n_rows)
            row_labels.append(r0 + 1)
            srow = []
            for bc in range(n_bin_cols):
                c0, c1 = bc * MINIMAP_BIN, min(bc * MINIMAP_BIN + MINIMAP_BIN, n_cols)
                cell_vals = [values[r][c] for r in range(r0, r1) for c in range(c0, c1)]
                if (br, bc) in corner_bins:
                    structural = "#"
                elif all(v is None for v in cell_vals):
                    structural = "."
                else:
                    structural = ":"
                if structural == ".":
                    code = ".----"
                else:
                    has_d = has_a = has_b = has_c = False
                    for v in cell_vals:
                        date, metric, entity, year = _cell_flags(v, metric_term, entity_term, year_term)
                        has_d = has_d or date
                        has_a = has_a or metric
                        has_b = has_b or entity
                        has_c = has_c or year
                        if has_d and has_a and has_b and has_c:
                            break
                    code = structural + ("D" if has_d else "-") + ("A" if has_a else "-") + ("B" if has_b else "-") + ("C" if has_c else "-")
                srow.append(code)
            symbol_rows.append(srow)
        col_labels = [bc * MINIMAP_BIN + 1 for bc in range(n_bin_cols)]

        minimap = _render_grid(symbol_rows, row_labels, col_labels)
        result = {
            "minimap": minimap,
            "legend": (
                "5-char bin code [structural][D][A][B][C]. structural: '.' blank bin  ':' content, "
                "not a corner  '#' block top-left corner. D/A/B/C: letter if a date-like/metric-match/"
                "entity-match/year-match cell exists anywhere in the bin, else '-'. Blank bin is "
                "always '.----'. (bin = 10x10 cells)"
            ),
            "scanned_rows": n_rows,
            "scanned_cols": n_cols,
        }
        if true_max_row > n_rows:
            result["note"] = f"{sheet} not fully scanned, {true_max_row - n_rows} rows remaining"
        return result

    def peek_ascii(self, sheet, r0, r1, c0, c1, metric_term=_UNSET, entity_term=_UNSET, year_term=_UNSET):
        metric_term = _resolve_term(metric_term, self._metric_term)
        entity_term = _resolve_term(entity_term, self._entity_term)
        year_term = _resolve_term(year_term, self._year_term)
        self._check_visible(sheet)
        r1c, c1c, note = _clamp_window(r0, r1, c0, c1, PEEK_ASCII_MAX_CELLS)
        ws = self.wb[sheet]
        rows = list(ws.iter_rows(min_row=r0, max_row=r1c, min_col=c0, max_col=c1c, values_only=True))
        symbol_rows = [[_cell_code(v, metric_term, entity_term, year_term) for v in row] for row in rows]
        grid = _render_grid(symbol_rows, list(range(r0, r1c + 1)), list(range(c0, c1c + 1)))
        result = {
            "grid": grid,
            "legend": ". blank  A metric-match  B entity-match  C year-match  D date-like  S string  N number  ? other (priority: A > B > C > D > S/N/?)",
        }
        if note:
            result["note"] = note
        return result

    def peek(self, sheet, r0, r1, c0, c1):
        self._check_visible(sheet)
        r1c, c1c, note = _clamp_window(r0, r1, c0, c1, PEEK_MAX_CELLS)
        ws = self.wb[sheet]
        rows = list(ws.iter_rows(min_row=r0, max_row=r1c, min_col=c0, max_col=c1c, values_only=True))
        result = {"rows": {r0 + i: {c0 + j: _round_display(v) for j, v in enumerate(row)} for i, row in enumerate(rows)}}
        error_count = sum(1 for row in rows for v in row if is_excel_error(v))
        if error_count:
            result["warning"] = (
                f"{error_count} cell(s) in this window are formula errors (#NAME?/#DIV/0!/etc.) — "
                f"likely broken external data links, not real data"
            )
        if note:
            result["note"] = note
        return result

    def _peek_axis(self, sheet, row, col, fixed_is_col):
        """Shared implementation for peek_col (fixed_is_col=True) and peek_row
        (fixed_is_col=False). wsnSearch4.md item 2 §2."""
        self._check_visible(sheet)
        values, n_rows, n_cols, true_max_row, true_max_col = self._scan(sheet)
        # The anchor's fixed axis (its own column for peek_col, its own row for
        # peek_row) indexes the scanned matrix directly, unwindowed — out-of-range
        # here would otherwise raise before any note could explain why.
        if fixed_is_col and not (1 <= col <= n_cols):
            return {"values": {}, "note": f"column {col} is outside the sheet's scanned extent (1-{n_cols})"}
        if not fixed_is_col and not (1 <= row <= n_rows):
            return {"values": {}, "note": f"row {row} is outside the sheet's scanned extent (1-{n_rows})"}
        entries = _merge_blocks(_detect_blocks(values, n_rows, n_cols))
        entry = _find_entry(entries, row, col)
        note = None
        if entry is None:
            if fixed_is_col:
                lo = max(1, row - PEEK_AXIS_WINDOW_RADIUS)
                hi = min(n_rows, row + PEEK_AXIS_WINDOW_RADIUS)
            else:
                lo = max(1, col - PEEK_AXIS_WINDOW_RADIUS)
                hi = min(n_cols, col + PEEK_AXIS_WINDOW_RADIUS)
            # Anchor coordinate itself beyond the scanned extent (e.g. col=9999 on a
            # 48-col sheet) makes lo>hi here — collapse to the nearest valid edge
            # instead of reporting a backwards range.
            if lo > hi:
                lo = hi
            note = f"no block detected, showing {lo}-{hi}"
        else:
            r0, r1 = entry["rows"]
            c0, c1 = entry["cols"]
            if fixed_is_col:
                span = r1 - r0 + 1
                if span <= 2 * PEEK_AXIS_WINDOW_RADIUS + 1:
                    lo, hi = r0, r1
                else:
                    lo = max(r0, row - PEEK_AXIS_WINDOW_RADIUS)
                    hi = min(r1, row + PEEK_AXIS_WINDOW_RADIUS)
                    note = f"block spans {span} rows, showing {lo}-{hi} centered on the requested cell"
            else:
                span = c1 - c0 + 1
                if span <= 2 * PEEK_AXIS_WINDOW_RADIUS + 1:
                    lo, hi = c0, c1
                else:
                    lo = max(c0, col - PEEK_AXIS_WINDOW_RADIUS)
                    hi = min(c1, col + PEEK_AXIS_WINDOW_RADIUS)
                    note = f"block spans {span} cols, showing {lo}-{hi} centered on the requested cell"

        values_out = {}
        if fixed_is_col:
            for r in range(lo, hi + 1):
                v = values[r - 1][col - 1]
                if v is not None:
                    values_out[r] = _round_display(v)
            anchor_index = row
        else:
            for cc in range(lo, hi + 1):
                v = values[row - 1][cc - 1]
                if v is not None:
                    values_out[cc] = _round_display(v)
            anchor_index = col

        values_out, trim_note = _trim_to_char_limit(values_out, anchor_index, PEEK_AXIS_CHAR_LIMIT)
        if trim_note:
            note = trim_note
        elif not values_out:
            axis_word = "column" if fixed_is_col else "row"
            note = f"{axis_word} {col if fixed_is_col else row} is blank throughout the {lo}-{hi} range checked"

        result = {"values": values_out}
        if note:
            result["note"] = note
        return result

    def peek_col(self, sheet, row, col):
        return self._peek_axis(sheet, row, col, fixed_is_col=True)

    def peek_row(self, sheet, row, col):
        return self._peek_axis(sheet, row, col, fixed_is_col=False)

    def all_axis(self, sheet, kind, index):
        self._check_visible(sheet)
        ws = self.wb[sheet]
        note = None
        if kind == "row":
            true_max_col = ws.max_column
            vals = list(ws.iter_rows(min_row=index, max_row=index, min_col=1, max_col=true_max_col, values_only=True))[0]
        elif kind == "col":
            true_max_row = ws.max_row
            max_row = min(true_max_row, ROW_SCAN_BUDGET)
            vals = [row[0] for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=index, max_col=index, values_only=True)]
            if true_max_row > max_row:
                note = f"axis not fully scanned, {true_max_row - max_row} rows remaining"
        else:
            return {"error": "kind must be 'row' or 'col'"}

        values = {i + 1: _round_display(v) for i, v in enumerate(vals) if v is not None}
        if len(values) > AXIS_VALUE_CAP:
            total_found = len(values)
            keep = sorted(values.keys())[:AXIS_VALUE_CAP]
            values = {k: values[k] for k in keep}
            # Actual final constraint on what's returned wins over the scan-depth note,
            # same precedent as item 2 §2 step 6 (char net overwrites span note).
            note = f"too many values! truncated to {AXIS_VALUE_CAP} of {total_found} found"

        result = {"values": values}
        warnings = []
        if kind == "col" and values:
            numeric_frac = sum(isinstance(v, (int, float)) for v in values.values()) / len(values)
            if numeric_frac > 0.5:
                warnings.append("majority of values are numeric — this may be a data column, not a label column")
        error_count = sum(1 for v in values.values() if is_excel_error(v))
        if error_count:
            warnings.append(
                f"{error_count} value(s) on this axis are formula errors (#NAME?/#DIV/0!/etc.) — "
                f"likely broken external data links, not real data"
            )
        if warnings:
            result["warnings"] = warnings
        if note:
            result["note"] = note
        return result

    def read_cell(self, sheet, row, col):
        self._check_visible(sheet)
        ws = self.wb[sheet]
        v = list(ws.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col, values_only=True))[0][0]
        result = {"value": _round_display(v)}
        if is_excel_error(v):
            result["error"] = True
            result["note"] = "this is a formula error, likely a broken external data link (e.g. disconnected Bloomberg add-in), not real data"
        return result

    def check_axes(self, sheet, tags, rows, cols):
        """wsnSearch5.md item 3a. tags/rows/cols are copied verbatim off one map_block output
        row — no independent block re-detection. Renders the whole block, tag-centered and
        clamped only if over PEEK_ASCII_MAX_CELLS: tagged rows/cols show literal values,
        everything else shows peek_ascii's type flag. Reads metric_term/entity_term/year_term
        from session state (set_terms) unconditionally — no term params of its own."""
        self._check_visible(sheet)
        r0, r1 = _parse_range(rows)
        c0, c1 = _parse_range(cols)
        values, n_rows, n_cols, true_max_row, true_max_col = self._scan(sheet)
        r1 = min(r1, n_rows)
        c1 = min(c1, n_cols)
        metric_term, entity_term, year_term = self._metric_term, self._entity_term, self._year_term

        parsed = [_parse_tag(t) for t in tags]
        row_tags = sorted({idx for idx, orient in parsed if orient == "r"})
        col_tags = sorted({idx for idx, orient in parsed if orient == "c"})
        if not row_tags and not col_tags:
            return {"grid": "", "note": "no valid tags parsed"}

        def _is_header_like(v):
            return _cell_code(v, metric_term, entity_term, year_term) in ("D", "S")

        notes = []
        extra_rows = []
        for idx in row_tags:
            above = idx - 1
            if above < r0 or above in row_tags or above in extra_rows:
                continue
            line = [values[above - 1][cc - 1] for cc in range(c0, c1 + 1)]
            if any(v is not None for v in line) and any(_is_header_like(v) for v in line):
                extra_rows.append(above)
                notes.append(f"row {above} included above requested row {idx} — header content detected, not itself a requested tag")

        extra_cols = []
        for idx in col_tags:
            left = idx - 1
            if left < c0 or left in col_tags or left in extra_cols:
                continue
            line = [values[r - 1][left - 1] for r in range(r0, r1 + 1)]
            if any(v is not None for v in line) and any(_is_header_like(v) for v in line):
                extra_cols.append(left)
                notes.append(f"col {left} included left of requested col {idx} — header content detected, not itself a requested tag")

        all_row_lines = sorted(set(row_tags) | set(extra_rows))
        all_col_lines = sorted(set(col_tags) | set(extra_cols))

        # Tag-centered bounding box (wsnSearch5.md item 3a §2 step 7 "Concrete fix"): an
        # orientation with no tags at all defaults to the block's full range for that axis —
        # "always whole block", not span-by-tag-composition.
        box_r0, box_r1 = (min(all_row_lines), max(all_row_lines)) if all_row_lines else (r0, r1)
        box_c0, box_c1 = (min(all_col_lines), max(all_col_lines)) if all_col_lines else (c0, c1)

        win_r0 = max(r0, box_r0 - PEEK_AXIS_WINDOW_RADIUS)
        win_r1 = min(r1, box_r1 + PEEK_AXIS_WINDOW_RADIUS)
        win_c0 = max(c0, box_c0 - PEEK_AXIS_WINDOW_RADIUS)
        win_c1 = min(c1, box_c1 + PEEK_AXIS_WINDOW_RADIUS)

        if (win_r1 - win_r0 + 1) * (win_c1 - win_c0 + 1) > PEEK_ASCII_MAX_CELLS:
            win_r1, win_c1, clamp_note = _clamp_window(win_r0, win_r1, win_c0, win_c1, PEEK_ASCII_MAX_CELLS)
            if clamp_note:
                notes.append(clamp_note)

        row_line_set = set(all_row_lines)
        col_line_set = set(all_col_lines)
        symbol_rows = []
        for r in range(win_r0, win_r1 + 1):
            row_is_tagged = r in row_line_set
            srow = []
            for cc in range(win_c0, win_c1 + 1):
                v = values[r - 1][cc - 1]
                if row_is_tagged or cc in col_line_set:
                    srow.append(_round_display(v) if v is not None else ".")
                else:
                    srow.append(_cell_code(v, metric_term, entity_term, year_term))
            symbol_rows.append(srow)

        grid = _render_grid_var(symbol_rows, list(range(win_r0, win_r1 + 1)), list(range(win_c0, win_c1 + 1)))

        # Tagged-line values, restated as an explicit {index: value} dict per line — same shape
        # peek_axes already uses. The grid's own literal cells answer "does real data line up
        # under both axes" (a visual/pattern question); this answers "what is the value at column
        # N" without requiring a count across the grid's header to find column N in the first
        # place. Wide tagged columns (a long literal label, or several 20-char datetime columns)
        # spread the header's index numbers unevenly, which measurably caused exactly this
        # miscount in testing (wsnSearch5_testlog.md) — this listing is the fix, not a cosmetic
        # addition. Bounded by the same rendered window as the grid, so it can't blow past
        # PEEK_ASCII_MAX_CELLS on its own.
        tagged_lines = []
        for r in sorted(row_line_set):
            d = {}
            for cc in range(win_c0, win_c1 + 1):
                v = values[r - 1][cc - 1]
                if v is not None:
                    d[cc] = _round_display(v)
            tagged_lines.append(f"{r}r: {d}")
        for cc in sorted(col_line_set):
            d = {}
            for r in range(win_r0, win_r1 + 1):
                v = values[r - 1][cc - 1]
                if v is not None:
                    d[r] = _round_display(v)
            tagged_lines.append(f"{cc}c: {d}")

        result = {
            "grid": grid,
            "tagged_values": "\n".join(tagged_lines),
            "legend": (
                "tagged rows/cols show their real value (3 significant figures) in the grid; every "
                "other cell shows peek_ascii's type flag — . blank  A metric-match  B entity-match  "
                "C year-match  D date-like  S string  N number  ? other. tagged_values restates "
                "each tagged row/col's own values as an explicit {column_or_row_index: value} dict — "
                "use it to read the actual value at a given index, don't count grid columns by eye."
            ),
        }
        if notes:
            result["note"] = "; ".join(notes)
        return result

    def peek_axes(self, sheet, tags, rows, cols):
        """wsnSearch5.md item 3b. Same sourcing convention as check_axes, but rows/cols only
        bound the scan range per tag — no rendered window, no flagging, no term dependency.
        Sparse {index: value} dict per tag, independent AXIS_VALUE_CAP truncation each."""
        self._check_visible(sheet)
        r0, r1 = _parse_range(rows)
        c0, c1 = _parse_range(cols)
        values, n_rows, n_cols, true_max_row, true_max_col = self._scan(sheet)
        r1 = min(r1, n_rows)
        c1 = min(c1, n_cols)

        lines = []
        for t in tags:
            idx, orient = _parse_tag(t)
            line_values = {}
            if orient == "r":
                for cc in range(c0, c1 + 1):
                    v = values[idx - 1][cc - 1]
                    if v is not None:
                        line_values[cc] = _round_display(v)
            else:
                for r in range(r0, r1 + 1):
                    v = values[r - 1][idx - 1]
                    if v is not None:
                        line_values[r] = _round_display(v)
            note = None
            if len(line_values) > AXIS_VALUE_CAP:
                total_found = len(line_values)
                keep = sorted(line_values.keys())[:AXIS_VALUE_CAP]
                line_values = {k: line_values[k] for k in keep}
                note = f"too many values! truncated to {AXIS_VALUE_CAP} of {total_found} found"
            lines.append((t, line_values, note))

        text = "\n".join(
            f"{t}: {vals}" + (f"  note: {note}" if note else "")
            for t, vals, note in lines
        )
        return {"axes": text}


def call_deepseek(api_key, model, messages):
    resp = requests.post(
        DEEPSEEK_BASE_URL,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"model": model, "messages": messages, "tools": TOOLS},
        timeout=120,
    )
    resp.raise_for_status()
    return resp.json()


def run(task, model, xlsx_path, max_tool_calls=MAX_TOOL_CALLS):
    env = load_env(ENV_PATH)
    api_key = env.get("DEEPSEEK_API_KEY")
    if not api_key:
        sys.exit("DEEPSEEK_API_KEY not found in .env")

    with open(PROTOCOL_PATH) as f:
        protocol = f.read()

    workbook = Workbook(xlsx_path)
    dispatch = {
        "list_sheets": lambda **kw: workbook.list_sheets(**kw),
        "get_dims": lambda **kw: workbook.get_dims(**kw),
        "map_block": lambda **kw: workbook.map_block(**kw),
        "map_bin": lambda **kw: workbook.map_bin(**kw),
        "peek_ascii": lambda **kw: workbook.peek_ascii(**kw),
        "peek": lambda **kw: workbook.peek(**kw),
        "peek_col": lambda **kw: workbook.peek_col(**kw),
        "peek_row": lambda **kw: workbook.peek_row(**kw),
        "all_axis": lambda **kw: workbook.all_axis(**kw),
        "read_cell": lambda **kw: workbook.read_cell(**kw),
        "set_terms": lambda **kw: workbook.set_terms(**kw),
        "check_axes": lambda **kw: workbook.check_axes(**kw),
        "peek_axes": lambda **kw: workbook.peek_axes(**kw),
    }

    system_prompt = "Follow this protocol exactly. Use only the tools provided.\n\n" + protocol
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": task},
    ]

    seen_calls = {}
    tool_call_count = 0
    # prompt_cache_hit_tokens/prompt_cache_miss_tokens come straight from DeepSeek's
    # response usage object (disk cache is automatic, no client setup) and sum to
    # prompt_tokens for that turn — no need to derive hit/miss from turn-to-turn deltas.
    usage_totals = {
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "total_tokens": 0,
        "prompt_cache_hit_tokens": 0,
        "prompt_cache_miss_tokens": 0,
    }
    api_turns = 0

    while True:
        data = call_deepseek(api_key, model, messages)
        api_turns += 1
        usage = data.get("usage", {})
        for k in usage_totals:
            usage_totals[k] += usage.get(k, 0)

        message = data["choices"][0]["message"]
        messages.append(message)

        tool_calls = message.get("tool_calls")
        if not tool_calls:
            print(message.get("content", ""))
            break

        budget_exhausted = False
        for tc in tool_calls:
            name = tc["function"]["name"]
            args = json.loads(tc["function"]["arguments"] or "{}")

            if tool_call_count >= max_tool_calls:
                budget_exhausted = True
                output = {"error": "tool call budget exhausted, stop and answer with what you have"}
            else:
                key = (name, json.dumps(args, sort_keys=True))
                if key in seen_calls:
                    output = {"note": "duplicate call — identical to an earlier call this session, reuse that result instead of re-reading"}
                else:
                    tool_call_count += 1
                    try:
                        output = dispatch[name](**args)
                    except Exception as e:
                        output = {"error": str(e)}
                    seen_calls[key] = True

            output_str = json.dumps(output, default=str)
            if len(output_str) > MAX_TOOL_OUTPUT_CHARS:
                output_str = output_str[:MAX_TOOL_OUTPUT_CHARS] + f"... [truncated, {len(output_str)} chars total]"

            print(f"\n[{tool_call_count}/{max_tool_calls}] {name}({args}) ->\n{output_str}")

            messages.append({"role": "tool", "tool_call_id": tc["id"], "content": output_str})

        if budget_exhausted:
            print("\n[stopped: tool call budget exhausted]")
            break

    est_cost = (
        usage_totals["prompt_cache_hit_tokens"] / 1_000_000 * PRICE_PER_MTOK_INPUT_HIT
        + usage_totals["prompt_cache_miss_tokens"] / 1_000_000 * PRICE_PER_MTOK_INPUT_MISS
        + usage_totals["completion_tokens"] / 1_000_000 * PRICE_PER_MTOK_OUTPUT
    )
    # hit+miss should equal prompt_tokens; if a turn's response omitted the cache fields
    # (e.g. an error response), they undercount and est_cost understates that turn's input cost.
    print(
        f"\n--- usage: {api_turns} API turns, {tool_call_count} tool calls, "
        f"{usage_totals['prompt_tokens']} prompt tokens "
        f"({usage_totals['prompt_cache_hit_tokens']} cache hit, "
        f"{usage_totals['prompt_cache_miss_tokens']} cache miss), "
        f"{usage_totals['completion_tokens']} completion tokens, "
        f"{usage_totals['total_tokens']} total tokens, ~${est_cost:.4f} est. cost ---"
    )


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit('usage: python3 harness.py "<task>" <xlsx_filename> [model] [max_tool_calls]')
    task_arg = sys.argv[1]
    xlsx_arg = os.path.join(HARNESS_DIR, sys.argv[2])
    model_arg = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_MODEL
    max_calls_arg = int(sys.argv[4]) if len(sys.argv) > 4 else MAX_TOOL_CALLS
    run(task_arg, model_arg, xlsx_arg, max_calls_arg)
