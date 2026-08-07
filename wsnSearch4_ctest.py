#!/usr/bin/env python3
"""
date created: 2026-08-08
date updated: 2026-08-08
date surroundings last checked: 2026-08-08

Throwaway practicality test for wsnSearch4.md Item 2's C signal (year triangulation).
Not part of harness.py / dateregex.py. Column-scans real xlsx files at the answer cell's
column (D signal, reused from dateregex.py) to find header candidates, then runs a fresh
is_year_match (C signal, per wsnSearch4.md item 2 section 3) against them. Two named
exclusion patterns (from_yyyy_q, partial cjk_year_month_range) get a targeted regex search
of their source file instead of a guessed coordinate.
"""
import datetime
import re
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from dateregex import is_date_like

BASE = Path(__file__).parent

# ---- C signal: is_year_match, per wsnSearch4.md item 2 section 3 ----
# 4-digit patterns: existing ones (gaining capture groups) + 3 new ones.
_C_4DIGIT = [
    ("yyyy", re.compile(r"(\d{4})")),
    ("yyyy_e", re.compile(r"(\d{4})'?e")),
    ("fy_yyyy", re.compile(r"fy(\d{4})")),
    ("end_yyyy", re.compile(r"end-(\d{4})e?")),
    ("cjk_year_end", re.compile(r"(\d{4})年末")),
    ("ye_yyyy", re.compile(r"ye\s?(\d{4})")),
    ("cy_yyyy", re.compile(r"cy(\d{4})")),
    ("yyyy_a", re.compile(r"(\d{4})a")),
]
# 2-digit fiscal-embedded patterns, compared against target_year % 100.
_C_2DIGIT = [
    ("fy3_yy", re.compile(r"fy3/(\d{2})e?")),
    ("q_yy", re.compile(r"[1-4]q(\d{2})")),
    ("q_yy_e", re.compile(r"[1-4]q(\d{2})e")),
    ("qfy_yy", re.compile(r"[1-4]qfy(\d{2})e?")),
    ("h_yy", re.compile(r"[1-2]h(\d{2})")),
    ("m_yy", re.compile(r"\d{1,2}m(\d{2})")),
    ("range_yyyy_yy", re.compile(r"\d{4}-(\d{2})")),
    ("range_yyyy_yy_e", re.compile(r"\d{4}-(\d{2})\s?e")),
    ("scenario_yy_e", re.compile(r"(?:bull|bear)\s'(\d{2})e")),
]
_CJK_YEAR_MONTH_RANGE = re.compile(r"(\d{4})年(\d{1,2})-(\d{1,2})月")
_FROM_YYYY_Q = re.compile(r"from\s(\d{4})\s[1-4]q")


def is_year_match(v, target_year):
    if isinstance(v, datetime.datetime):
        return v.year == target_year
    if isinstance(v, datetime.time):
        return False  # bare time-of-day carries no year; out of scope for C
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if float(v).is_integer():
            iv = int(v)
            if 1980 <= iv <= 2035:
                return iv == target_year
        return False
    if isinstance(v, str):
        s = v.strip().casefold()
        if not s:
            return False
        m = _CJK_YEAR_MONTH_RANGE.fullmatch(s)
        if m:
            y, m1, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return y == target_year and m1 == 1 and m2 == 12
        if _FROM_YYYY_Q.fullmatch(s):
            return False
        for _pid, regex in _C_4DIGIT:
            m = regex.fullmatch(s)
            if m:
                return int(m.group(1)) == target_year
        for _pid, regex in _C_2DIGIT:
            m = regex.fullmatch(s)
            if m:
                return int(m.group(1)) == target_year % 100
        return False
    return False


# ---- helpers ----
def load(fname):
    return openpyxl.load_workbook(BASE / fname, read_only=True, data_only=True)


def cell_ref_to_rc(ref):
    col_str, row = coordinate_from_string(ref)
    return row, column_index_from_string(col_str)


def col_scan(ws, col, max_row):
    """Every non-empty cell in one column, rows 1..max_row. NOT gated by D --
    C is evaluated independently, since dateregex.is_date_like has no numeric-cell
    path at all (confirmed below) and would silently hide those candidates from C."""
    out = []
    for i, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=col, max_col=col), start=1):
        v = row_cells[0].value
        if v is not None:
            out.append((i, v))
    return out


def row_neighbors(ws, row, col, radius=3):
    """Non-empty cells in the same row, columns col-radius..col+radius (excluding col itself)."""
    out = []
    lo = max(1, col - radius)
    hi = col + radius
    for row_cells in ws.iter_rows(min_row=row, max_row=row, min_col=lo, max_col=hi):
        for offset, cell in enumerate(row_cells):
            actual_col = lo + offset
            if actual_col == col:
                continue
            if cell.value is not None:
                out.append((actual_col, cell.value))
    return out


def targeted_search(ws, pattern, max_row=200, max_col=60):
    out = []
    for r, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1):
        for c, cell in enumerate(row_cells, start=1):
            if isinstance(cell.value, str):
                s = cell.value.strip().casefold()
                if s and pattern.fullmatch(s):
                    out.append((r, c, cell.value))
    return out


# ---- test rows, from testqs.md ----
TEST_ROWS = [
    dict(file="300450_Wuxi Lead_20210422_client.xlsx", sheet="Op & Non op", cell="L22",
         target_year=2022, pattern="numeric-typed yyyy", expect=True),
    dict(file="300450_Wuxi Lead_20210422_client.xlsx", sheet="Consolidate", cell="M46",
         target_year=2022, pattern="string yyyy_e (2022E)", expect=True),
    dict(file="300450_Wuxi Lead_20210422_client.xlsx", sheet="PL", cell="AA6",
         target_year=2020, pattern="half-year h_yy (1H20)", expect=True),
    dict(file="20220404_Parade_BOE.xlsx", sheet="Shipment", cell="G10",
         target_year=2018, pattern="quarter q_yy (1Q18)", expect=True),
    dict(file="[Quarter added]Advantest 6857 JP BOE.xlsx", sheet="Advantest BOE", cell="AI86",
         target_year=2022, pattern="range_yyyy_yyyy (2022-2025E) -- expected MISS, out of scope",
         expect=False),
    dict(file="[Quarter added]Advantest 6857 JP BOE.xlsx", sheet="Advantest BOE", cell="AC86",
         target_year=2022, pattern="fy3_yy (FY3/22)", expect=True),
    dict(file="[Quarter added]Advantest 6857 JP BOE.xlsx", sheet="Advantest BOE", cell="D8",
         target_year=2020, pattern="qfy_yy / q_yy (4QFY20 / 1Q20)", expect=True),
    dict(file="20220204_WaferSupplyDemandModel.xlsx", sheet="PriceAssumption", cell="F37",
         target_year=2018, pattern="cjk_year_end (2018年末)", expect=True),
    dict(file="99-00 Study Raw.xlsx", sheet="Existing", cell="G5",
         target_year=1998, pattern="native datetime (Feb 1998)", expect=True),
    dict(file="99-00 Study Raw.xlsx", sheet="Existing", cell="H603",
         target_year=1998, pattern="native datetime (Mar 1998, row 603)", expect=True),
    dict(file="2.元戎启行资金消耗统计 2019.2-2021.4.xlsx", sheet="元戎启行资金消耗统计", cell="C4",
         target_year=2019, pattern="cjk_year_month_range full (2019年1-12月)", expect=True),
    dict(file="2.元戎启行资金消耗统计 2019.2-2021.4.xlsx", sheet="元戎启行资金消耗统计", cell="E5",
         target_year=2020, pattern="cjk_year_month single (2020年2月) -- not in C spec, expect MISS",
         expect=False),
    dict(file="2.元戎启行资金消耗统计 2019.2-2021.4.xlsx", sheet="元戎启行资金消耗统计", cell="S6",
         target_year=2021, pattern="cjk_year_month single (2021年4月) -- not in C spec, expect MISS",
         expect=False),
]


def run_main_tests():
    rows = []
    wb_cache = {}
    for t in TEST_ROWS:
        if t["file"] not in wb_cache:
            wb_cache[t["file"]] = load(t["file"])
        wb = wb_cache[t["file"]]
        ws = wb[t["sheet"]]
        row, col = cell_ref_to_rc(t["cell"])
        candidates = col_scan(ws, col, row)
        any_true = False
        detail = []
        for r, v in candidates:
            c = is_year_match(v, t["target_year"])
            d = is_date_like(v)
            any_true = any_true or c
            detail.append((r, v, c, d))
        status = "PASS" if any_true == t["expect"] else "FAIL"
        rows.append((t, candidates, detail, any_true, status))
    return rows


def print_main_report(rows):
    print("=" * 100)
    print("MAIN TEST ROWS")
    print("=" * 100)
    for t, candidates, detail, any_true, status in rows:
        print(f"\n[{status}] {t['file']} | {t['sheet']} | {t['cell']} | target_year={t['target_year']}")
        print(f"  pattern: {t['pattern']}")
        print(f"  non-empty cells in column, rows 1..{cell_ref_to_rc(t['cell'])[0]}: {len(candidates)}")
        for r, v, c, d in detail:
            if not (c or d):
                continue
            print(f"    row {r:>4}: C={c!s:<5} D={d!s:<5} value={v!r}")
        print(f"  any C match: {any_true}  expected: {t['expect']}")

    # neighbor decoys for a subset with a clear positive header row found
    print("\n" + "=" * 100)
    print("NEIGHBOR-COLUMN DECOYS (same header row, +/-3 columns)")
    print("=" * 100)
    for t, candidates, detail, any_true, status in rows:
        if not t["expect"]:
            continue
        true_rows = [r for r, v, c, d in detail if c]
        if not true_rows:
            continue
        header_row = true_rows[0]
        wb = wb_cache_global[t["file"]]
        ws = wb[t["sheet"]]
        _, col = cell_ref_to_rc(t["cell"])
        neighbors = row_neighbors(ws, header_row, col)
        print(f"\n{t['file']} | {t['sheet']} | header row {header_row}, target column {col}")
        if not neighbors:
            print("    (no neighbor cells within +/-3 columns)")
        for ncol, v in neighbors:
            c = is_year_match(v, t["target_year"])
            print(f"    col {ncol:>4}: C={c}  value={v!r}")


wb_cache_global = {}


def run_exclusion_tests():
    print("\n" + "=" * 100)
    print("TARGETED EXCLUSION-PATTERN SEARCH (live)")
    print("=" * 100)

    wb = load("20220204_WaferSupplyDemandModel.xlsx")
    found_any = False
    for sheet_name in wb.sheetnames:
        if "wafer" not in sheet_name.casefold() and "supply" not in sheet_name.casefold():
            continue
        ws = wb[sheet_name]
        hits = targeted_search(ws, _FROM_YYYY_Q, max_row=200, max_col=60)
        for r, c, v in hits:
            found_any = True
            result = is_year_match(v, 2016)
            print(f"  from_yyyy_q live hit: {sheet_name}!R{r}C{c} = {v!r}  is_year_match(2016)={result}  expected=False")
    if not found_any:
        print("  from_yyyy_q: no live match found in scanned bound (rows 1-200, cols 1-60) across wafer/supply sheets")

    wb2 = load("2.元戎启行资金消耗统计 2019.2-2021.4.xlsx")
    ws2 = wb2["元戎启行资金消耗统计"]
    hits = targeted_search(ws2, _CJK_YEAR_MONTH_RANGE, max_row=200, max_col=60)
    partial_found = False
    for r, c, v in hits:
        s = v.strip().casefold()
        m = _CJK_YEAR_MONTH_RANGE.fullmatch(s)
        m1, m2 = int(m.group(2)), int(m.group(3))
        if (m1, m2) != (1, 12):
            partial_found = True
            result = is_year_match(v, int(m.group(1)))
            print(f"  cjk_year_month_range PARTIAL live hit: R{r}C{c} = {v!r}  is_year_match={result}  expected=False")
    if not partial_found:
        print("  cjk_year_month_range partial: no live partial-range cell found in scanned bound")
        synthetic = "2019年3-6月"
        result = is_year_match(synthetic, 2019)
        print(f"  [SYNTHETIC, gate unit test only] is_year_match({synthetic!r}, 2019) = {result}  expected=False")


if __name__ == "__main__":
    rows = run_main_tests()
    wb_cache_global = {}
    for t in TEST_ROWS:
        if t["file"] not in wb_cache_global:
            wb_cache_global[t["file"]] = load(t["file"])
    print_main_report(rows)
    run_exclusion_tests()

    total = len(rows)
    passed = sum(1 for *_r, status in rows if status == "PASS")
    print("\n" + "=" * 100)
    print(f"SUMMARY: {passed}/{total} main rows matched expectation")
    print("=" * 100)
